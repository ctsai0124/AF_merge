from flask import Flask, request, jsonify, render_template, send_file
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io, os, json, re, tempfile, subprocess
from docx import Document
from docx.oxml.ns import qn

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

SIMPLE_COLS = ['清冊序號', '姓名', '薪俸表別', '總金額', '支領數額',
               '專業加給表別', '總金額.1', '支領數額.1',
               '職務加給表別', '總金額.2', '支領數額.2']

SIMPLE_REGION_COLS = ['清冊序號', '姓名', '薪俸表別', '總金額', '支領數額',
                      '專業加給表別', '總金額.1', '支領數額.1',
                      '職務加給表別', '總金額.2', '支領數額.2',
                      '地域加給表別', '總金額.3', '支領數額.3']

NUM_COLS = {'清冊序號', '總金額', '支領數額', '待遇差額', '補發金額', '增支',
            '總金額.1', '支領數額.1', '待遇差額.1', '補發金額.1',
            '總金額.2', '支領數額.2', '待遇差額.2', '補發金額.2',
            '總金額.3', '支領數額.3', '待遇差額.3', '補發金額.3'}

# ── PDF 比對：見 paycheck.py ──
import paycheck
import threading, time

# ── OCR 工作佇列（供 Mac 端 Vision 辨識使用）──────────────
import base64, uuid
_ocr_lock = threading.Lock()
_ocr_jobs = {}          # job_id -> dict
OCR_JOB_TTL = 600       # 工作逾時秒數

# 輪詢節奏：有人在用時密集詢問，閒置時放慢，減少無謂請求
ACTIVE_WINDOW = 600     # 最後一次操作起算的活躍期（秒）
POLL_ACTIVE = 3         # 活躍期的輪詢間隔
POLL_IDLE = 60          # 閒置期的輪詢間隔
_last_active = 0.0


def touch_active():
    """記錄有使用者正在操作"""
    global _last_active
    _last_active = time.time()


def next_poll_sec():
    """依目前是否活躍，回傳建議的輪詢間隔"""
    if _ocr_jobs:
        return POLL_ACTIVE
    return POLL_ACTIVE if (time.time() - _last_active) < ACTIVE_WINDOW else POLL_IDLE


def _ocr_key():
    return os.environ.get('OCR_KEY', '').strip()


def _ocr_auth(req):
    k = _ocr_key()
    return bool(k) and req.headers.get('X-OCR-KEY', '') == k


def _ocr_gc():
    now = time.time()
    for jid in [j for j, v in _ocr_jobs.items() if now - v['created'] > OCR_JOB_TTL]:
        _ocr_jobs.pop(jid, None)


def ocr_enabled():
    return bool(_ocr_key())


# ── 排除設定（依學校記憶）────────────────────────────────
_ex_lock = threading.Lock()


def _ex_file():
    base = os.environ.get('DATA_DIR', '').strip() or app.root_path
    try:
        os.makedirs(base, exist_ok=True)
    except Exception:
        base = tempfile.gettempdir()
    return os.path.join(base, 'exclusions.json')


def load_exclusions(school=''):
    try:
        with open(_ex_file(), encoding='utf-8') as f:
            d = json.load(f)
    except Exception:
        d = {}
    e = d.get(school or '_default', {})
    return {'titles': e.get('titles', []), 'names': e.get('names', []),
            'use_default': e.get('use_default', True)}


def save_exclusions(school, titles, names, use_default=None):
    with _ex_lock:
        try:
            with open(_ex_file(), encoding='utf-8') as f:
                d = json.load(f)
        except Exception:
            d = {}
        prev = d.get(school or '_default', {})
        d[school or '_default'] = {
            'titles': sorted({t.strip() for t in titles if t.strip()}),
            'names': sorted({n.strip() for n in names if n.strip()}),
            'use_default': prev.get('use_default', True) if use_default is None else bool(use_default),
        }
        try:
            with open(_ex_file(), 'w', encoding='utf-8') as f:
                json.dump(d, f, ensure_ascii=False)
        except Exception as e:
            app.logger.warning(f'排除設定寫入失敗：{e}')
        return d[school or '_default']


def school_key(af_name):
    sn1, _ = parse_af_filename(af_name or '')
    return sn1 or (af_name or '_default')


# ── 職稱觀察統計（跨檔案累積，用於建議排除）────────────────
_stats_lock = threading.Lock()


def _stats_file():
    base = os.environ.get('DATA_DIR', '').strip() or app.root_path
    try:
        os.makedirs(base, exist_ok=True)
    except Exception:
        base = tempfile.gettempdir()
    return os.path.join(base, 'title_stats.json')


def load_title_stats():
    try:
        with open(_stats_file(), encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def merge_title_stats(obs):
    """把本次觀察併入累積統計"""
    with _stats_lock:
        d = load_title_stats()
        for t, v in (obs or {}).items():
            e = d.setdefault(t, {'total': 0, 'in_af': 0, 'files': 0})
            e['total'] += v['total']
            e['in_af'] += v['in_af']
            e['files'] += 1
        try:
            with open(_stats_file(), 'w', encoding='utf-8') as f:
                json.dump(d, f, ensure_ascii=False)
        except Exception as e:
            app.logger.warning(f'職稱統計寫入失敗：{e}')
        return d


# ── 使用次數計數器 ────────────────────────────────────────
_counter_lock = threading.Lock()


def _counter_file():
    """計數檔位置：優先用 DATA_DIR（Volume），不可寫則退回專案目錄／暫存"""
    for base in (os.environ.get('DATA_DIR', '').strip(), app.root_path, tempfile.gettempdir()):
        if not base:
            continue
        try:
            os.makedirs(base, exist_ok=True)
            probe = os.path.join(base, '.write_test')
            with open(probe, 'w') as f:
                f.write('1')
            os.remove(probe)
            return os.path.join(base, 'counter.json')
        except Exception:
            continue
    return os.path.join(tempfile.gettempdir(), 'counter.json')


def counter_diag():
    """診斷資訊：確認計數檔實際寫到哪裡"""
    env = os.environ.get('DATA_DIR', '')
    path = _counter_file()
    return {
        'DATA_DIR環境變數': env if env else '（未設定）',
        '實際使用路徑': path,
        '是否寫入Volume': bool(env.strip()) and path.startswith(env.strip()),
        '檔案是否存在': os.path.exists(path),
    }


def load_counter():
    try:
        with open(_counter_file(), encoding='utf-8') as f:
            d = json.load(f)
    except Exception:
        d = {}
    return {'visits': d.get('visits', 0),
            'sorts': d.get('sorts', 0),
            'compares': d.get('compares', 0)}


def bump_counter(key):
    """累加某項計數，回傳更新後的完整計數"""
    with _counter_lock:
        d = load_counter()
        if key in d:
            d[key] += 1
        try:
            with open(_counter_file(), 'w', encoding='utf-8') as f:
                json.dump(d, f)
        except Exception as e:
            app.logger.warning(f'計數寫入失敗：{e}')
        return d


# ── 原有功能 ──────────────────────────────────────────────

def load_school_df():
    path = os.path.join(app.root_path, 'static', 'school.xlsx')
    if not os.path.exists(path):
        return pd.DataFrame(columns=['sn1', 'sn2', 'school'])
    return pd.read_excel(path, engine='openpyxl', dtype=str)


def parse_af_filename(filename):
    base = os.path.splitext(filename)[0]
    parts = base.split('_')
    sn1 = ''
    yearmonth = ''
    if len(parts) >= 2:
        sn1 = parts[-2]
    if len(parts) >= 1:
        ts = parts[-1]
        if len(ts) >= 5:
            yearmonth = ts[:5]
    return sn1, yearmonth


def lookup_school(sn1):
    df = load_school_df()
    if df.empty or not sn1:
        return '', ''
    row = df[df['sn1'].str.strip() == sn1.strip()]
    if row.empty:
        return '', ''
    return row.iloc[0]['school'], row.iloc[0]['sn2']


def read_sheet(file_bytes, filename, sheet_name):
    ext = os.path.splitext(filename)[1].lower()
    buf = io.BytesIO(file_bytes)
    engine = 'xlrd' if ext == '.xls' else 'openpyxl'
    return pd.read_excel(buf, sheet_name=sheet_name, engine=engine, dtype=str, header=0)


def build_excel(data, columns):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '排序結果'

    hdr_fill = PatternFill('solid', start_color='1a3a5c')
    hdr_font = Font(bold=True, color='FFFFFF', name='Microsoft JhengHei', size=11)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill('solid', start_color='EEF3F9')
    data_font = Font(name='Microsoft JhengHei', size=10)

    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border

    for ri, row in enumerate(data, 2):
        row_fill = alt_fill if ri % 2 == 0 else None
        for ci, col in enumerate(columns, 1):
            val = row.get(col, '')
            try:
                if col in NUM_COLS and val != '':
                    val = int(float(val))
            except (ValueError, TypeError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.border = border
            cell.alignment = center if col != '姓名' else left
            if row_fill:
                cell.fill = row_fill

    col_widths = {'清冊序號': 9, '姓名': 10, '薪俸表別': 12, '專業加給表別': 14,
                  '職務加給表別': 14, '地域加給表別': 14, '增支': 8,
                  '總金額': 10, '支領數額': 10, '待遇差額': 10, '補發金額': 10,
                  '總金額.1': 10, '支領數額.1': 10, '待遇差額.1': 10, '補發金額.1': 10,
                  '總金額.2': 10, '支領數額.2': 10, '待遇差額.2': 10, '補發金額.2': 10,
                  '總金額.3': 10, '支領數額.3': 10}
    for ci, col in enumerate(columns, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = col_widths.get(col, 11)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def replace_in_element(element, old, new):
    for para in element.findall('.//' + qn('w:p')):
        runs = para.findall(qn('w:r'))
        texts = []
        for r in runs:
            t = r.find(qn('w:t'))
            texts.append(t.text if t is not None else '')
        full = ''.join(texts)
        if old in full:
            new_full = full.replace(old, new)
            assigned = False
            for r in runs:
                t = r.find(qn('w:t'))
                if t is not None:
                    if not assigned:
                        t.text = new_full
                        assigned = True
                    else:
                        t.text = ''


def build_audit_docx(school_name, sn2, yearmonth):
    template_path = os.path.join(app.root_path, 'static', '稽核表-.docx')
    doc = Document(template_path)
    replace_in_element(doc.element.body, '<學校名稱>', school_name)
    replace_in_element(doc.element.body, '<年月份>', yearmonth)
    for txbx in doc.element.body.findall('.//' + qn('w:txbxContent')):
        para_texts = ''.join(t.text or '' for t in txbx.findall('.//' + qn('w:t')))
        if '<編號>' in para_texts:
            replace_in_element(txbx, '<編號>', sn2)
    out = io.BytesIO()
    doc.save(out)
    out.seek(0)
    return out

# ── 路由 ──────────────────────────────────────────────────

@app.route('/')
def index():
    bump_counter('visits')
    touch_active()
    return render_template('index.html')


@app.route('/stats')
def stats():
    return jsonify(load_counter())


@app.route('/stats/diag')
def stats_diag():
    d = counter_diag()
    d['目前計數'] = load_counter()
    return jsonify(d)


@app.route('/download-template')
def download_template():
    path = os.path.join(app.root_path, 'static', '固定清冊範例.xlsx')
    return send_file(path, as_attachment=True, download_name='固定清冊範例.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/download-standalone')
def download_standalone():
    path = os.path.join(app.root_path, 'static', 'AF欄位調整工具_單機d槽版.xlsm')
    return send_file(path, as_attachment=True, download_name='AF欄位調整工具_單機d槽版.xlsm',
                     mimetype='application/vnd.ms-excel.sheet.macroEnabled.12')


@app.route('/process', methods=['POST'])
def process():
    if 'roster' not in request.files or 'af' not in request.files:
        return jsonify({'error': '請上傳兩個檔案'}), 400

    roster_f = request.files['roster']
    af_f = request.files['af']
    roster_bytes = roster_f.read()
    af_bytes = af_f.read()

    try:
        roster_df = read_sheet(roster_bytes, roster_f.filename, 'input')
        roster_df.columns = [str(c).strip() for c in roster_df.columns]
        roster_df = roster_df[['序號', '姓名']].copy()
        roster_df['姓名'] = roster_df['姓名'].str.strip()
        roster_df['序號'] = pd.to_numeric(roster_df['序號'], errors='coerce')
        roster_df = roster_df.dropna(subset=['序號', '姓名'])
        roster_df = roster_df[roster_df['姓名'] != '']
        roster_df = roster_df.sort_values('序號')
        dup_roster = roster_df[roster_df.duplicated('姓名', keep=False)]['姓名'].unique().tolist()
        ordered_names = roster_df['姓名'].tolist()
        name_to_seq = {row['姓名']: int(row['序號']) for _, row in roster_df.iterrows()}

        af_df = read_sheet(af_bytes, af_f.filename, 0)
        af_df.columns = [str(c).strip() for c in af_df.columns]
        af_df['姓名'] = af_df['姓名'].str.strip()

        from collections import defaultdict
        af_map = defaultdict(list)
        for _, row in af_df.iterrows():
            af_map[row['姓名']].append(row)

        sorted_rows = []
        not_found = []
        found = set()

        for name in ordered_names:
            if name in af_map:
                seq = name_to_seq[name]
                for row in af_map[name]:
                    r = row.copy()
                    r['清冊序號'] = seq
                    sorted_rows.append(r)
                found.add(name)
            else:
                not_found.append(name)

        extra_seq = len(ordered_names) + 1
        extra_names = []
        for _, row in af_df.iterrows():
            if row['姓名'] not in found:
                r = row.copy()
                r['清冊序號'] = extra_seq
                sorted_rows.append(r)
                if row['姓名'] not in extra_names:
                    extra_names.append(row['姓名'])

        result_df = pd.DataFrame(sorted_rows).reset_index(drop=True)

        cols = result_df.columns.tolist()
        if '清冊序號' in cols:
            cols.remove('清冊序號')
        cols = ['清冊序號'] + cols
        result_df = result_df[cols]

        warnings = []
        if not_found:
            warnings.append('清冊中以下人員在 AF 找不到對應：' + '、'.join(not_found))
        if extra_names:
            warnings.append('AF 中以下人員不在清冊內，已附加至末尾：' + '、'.join(extra_names))
        if dup_roster:
            warnings.append('清冊中以下姓名出現多次，請確認是重複登打或同名不同人：' + '、'.join(dup_roster))

        sn1, yearmonth = parse_af_filename(af_f.filename)
        school_name, sn2 = lookup_school(sn1)

        app.config['LAST_RESULT'] = result_df.fillna('').to_json(orient='records', force_ascii=False)
        app.config['LAST_COLUMNS'] = result_df.columns.tolist()
        app.config['LAST_SCHOOL'] = school_name
        app.config['LAST_SN2'] = sn2
        app.config['LAST_YEARMONTH'] = yearmonth

        counts = bump_counter('sorts')

        return jsonify({
            'success': True,
            'counts': counts,
            'preview': result_df.head(20).fillna('').to_dict(orient='records'),
            'columns': result_df.columns.tolist(),
            'total': len(result_df),
            'warnings': warnings,
            'school_name': school_name,
            'sn2': sn2,
            'yearmonth': yearmonth
        })

    except KeyError as e:
        return jsonify({'error': '找不到欄位或工作表：' + str(e) + '，請確認檔案格式與範例相符'}), 400
    except Exception as e:
        return jsonify({'error': '處理錯誤：' + str(e)}), 500


# ── 新增：薪資清冊 PDF 比對 ────────────────────────────────

@app.route('/compare-pdf', methods=['POST'])
def compare_pdf():
    """薪資清冊 PDF 與 AF 比對；掃描檔則回傳手動輸入表格"""
    if 'salary_pdf' not in request.files:
        return jsonify({'error': '請上傳薪資清冊 PDF'}), 400
    if 'af' not in request.files or not request.files['af'].filename:
        return jsonify({'error': '請上傳 AF 資料檔'}), 400

    touch_active()
    af_f = request.files['af']
    af_bytes, af_name = af_f.read(), af_f.filename
    pdf_bytes = request.files['salary_pdf'].read()

    try:
        af_records, af_warns = paycheck.load_af(af_bytes, af_name)
    except KeyError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': '讀取 AF 失敗：' + str(e)}), 500

    # ── 掃描圖檔 → 無法解析 ──
    if not paycheck.has_text_layer(pdf_bytes):
        job_id = None
        if ocr_enabled():
            # 先保留檔案，使用者按下按鈕才真正送辨識（逾時自動清除）
            job_id = uuid.uuid4().hex[:12]
            with _ocr_lock:
                _ocr_gc()
                _ocr_jobs[job_id] = {'status': 'held', 'created': time.time(),
                                     'pdf': pdf_bytes, 'af': af_bytes, 'af_name': af_name,
                                     'people': None, 'error': None}

        return jsonify({
            'error': '此 PDF 為掃描圖檔，無法自動比對',
            'code': 'scanned',
            'ocr_available': bool(job_id),
            'job_id': job_id,
            'title': '這份 PDF 是掃描圖檔，無法比對',
            'why': '整份文件是一張張圖片，裡面沒有任何可讀取的文字資料，'
                   '系統無法取得薪俸、加給等金額。',
            'how': [
                '請向製表單位索取由薪資系統「直接匯出」的 PDF。',
                '分辨方法：用滑鼠在 PDF 上拖曳，選得到文字的才可以用；'
                '掃描檔只會選到一整塊區域。',
                '請注意：把掃描檔用文字辨識軟體轉換過的檔案同樣不行，'
                '因為表格的欄列對應關係已經被打散。'
            ]
        }), 422

    # ── 有文字層 → 直接解析比對 ──
    try:
        pdf_people, layout = paycheck.parse_pdf(pdf_bytes)
        if not pdf_people:
            return jsonify({
                'error': '偵測不到表格結構，無法比對',
                'code': 'no_table',
                'title': '這份 PDF 有文字，但讀不到表格',
                'why': '檔案裡雖然有文字，但欄與列的對應關係已經消失，'
                       '系統無法判斷哪個金額屬於哪個人、哪個項目。',
                'how': [
                    '最常見的原因是：這份檔案原本是掃描圖，'
                    '後來用文字辨識軟體（OCR）轉過一次。',
                    '轉換過程只會把文字攤平成一長串，表格會被拆散，'
                    '所以看得到文字也無法使用。',
                    '請改用薪資系統「直接匯出」的原始 PDF。'
                ]}), 400

        skey = school_key(af_name)
        ex = load_exclusions(skey)
        af_names = {a['姓名'] for a in af_records}
        obs = paycheck.title_observations(pdf_people, af_names)
        stats = merge_title_stats(obs)
        learned, _ = paycheck.learned_suggestions(stats)
        if learned:
            auto = sorted({x['職稱'] for x in learned} | set(ex['titles']))
            if auto != sorted(ex['titles']):
                save_exclusions(skey, auto, ex['names'])
                ex = load_exclusions(skey)

        paycheck.mark_exclusions(
            pdf_people, ex['titles'], ex['names'], ex.get('use_default', True))

        paycheck.annotate_arith(pdf_people)
        bad_arith = [p['姓名'] for p in pdf_people if p.get('_arith_ok') is False]
        if bad_arith:
            af_warns.append('薪資清冊本身加總不符（各項相加 ≠ 應發金額）：'
                            + '、'.join(bad_arith) + '，請先確認清冊是否正確')

        out = paycheck.compare(pdf_people, af_records)
        out['success'] = True
        out['mode'] = 'auto'
        out['layout'] = {'vertical': '直式（一人一欄）',
                         'horizontal': '橫式（一人一列）'}.get(layout, layout)
        out['af_count'] = len(af_records)
        out['pdf_count'] = len(pdf_people)
        out['warnings'] = af_warns
        out['school_key'] = skey
        out['counts'] = bump_counter('compares')
        return jsonify(out)

    except Exception as e:
        return jsonify({'error': '比對錯誤：' + str(e)}), 500


@app.route('/settings/title-stats')
def settings_title_stats():
    """檢視累積的職稱觀察統計"""
    stats = load_title_stats()
    learned, wrong = paycheck.learned_suggestions(stats)
    rows = sorted(({'職稱': t, **v} for t, v in stats.items()),
                  key=lambda x: -x['total'])
    return jsonify({'stats': rows, 'learned': learned, 'wrong': wrong})


@app.route('/settings/exclusions', methods=['GET', 'POST'])
def settings_exclusions():
    """讀取／儲存某校的排除設定"""
    if request.method == 'GET':
        return jsonify(load_exclusions(request.args.get('school', '')))
    d = request.get_json(silent=True) or {}
    return jsonify(save_exclusions(d.get('school', ''),
                                   d.get('titles') or [], d.get('names') or [],
                                   d.get('use_default')))


@app.route('/ocr/start/<job_id>', methods=['POST'])
def ocr_start(job_id):
    """使用者按下按鈕後，才將保留中的檔案送入辨識佇列"""
    touch_active()
    with _ocr_lock:
        j = _ocr_jobs.get(job_id)
        if not j:
            return jsonify({'error': '檔案已逾時清除，請重新上傳'}), 404
        if j['status'] == 'held':
            j['status'] = 'pending'
            j['created'] = time.time()
    return jsonify({'ok': True, 'job_id': job_id})


@app.route('/ocr/submit-fixed', methods=['POST'])
def ocr_submit_fixed():
    """使用者更正可疑列後，合併比對"""
    d = request.get_json(silent=True) or {}
    jid = d.get('job_id')
    with _ocr_lock:
        j = _ocr_jobs.get(jid)
        if not j:
            return jsonify({'error': '此次辨識結果已逾時，請重新上傳'}), 404
        good, af_bytes, af_name = j.get('good', []), j['af'], j['af_name']
    try:
        af_records, af_warns = paycheck.load_af(af_bytes, af_name)
        skey = school_key(af_name)
        if d.get('remember_titles') or d.get('remember_names'):
            cur = load_exclusions(skey)
            save_exclusions(skey,
                            set(cur['titles']) | set(d.get('remember_titles') or []),
                            set(cur['names']) | set(d.get('remember_names') or []))
        out = paycheck.compare_with_fixed(good, d.get('rows') or [], af_records)
        out.update({'status': 'done', 'success': True, 'mode': 'ocr',
                    'layout': '掃描圖檔（文字辨識＋人工確認）',
                    'af_count': len(af_records),
                    'pdf_count': len(good) + len(d.get('rows') or []),
                    'warnings': af_warns, 'counts': bump_counter('compares')})
        with _ocr_lock:
            _ocr_jobs.pop(jid, None)
        return jsonify(out)
    except Exception as e:
        return jsonify({'error': '比對錯誤：' + str(e)}), 500


@app.route('/ocr/claim')
def ocr_claim():
    """Mac 端領取待辨識工作"""
    if not _ocr_auth(request):
        return jsonify({'error': 'unauthorized'}), 401
    with _ocr_lock:
        _ocr_gc()
        for jid, j in _ocr_jobs.items():
            if j['status'] == 'pending':
                j['status'] = 'processing'
                return jsonify({'job_id': jid,
                                'pdf_b64': base64.b64encode(j['pdf']).decode(),
                                'next_poll': POLL_ACTIVE})
    return jsonify({'job_id': None, 'next_poll': next_poll_sec()})


@app.route('/ocr/result', methods=['POST'])
def ocr_result():
    """Mac 端回傳辨識結果"""
    if not _ocr_auth(request):
        return jsonify({'error': 'unauthorized'}), 401
    d = request.get_json(silent=True) or {}
    jid = d.get('job_id')
    with _ocr_lock:
        j = _ocr_jobs.get(jid)
        if not j:
            return jsonify({'error': 'job not found'}), 404
        if d.get('error'):
            j['status'] = 'failed'
            j['error'] = d['error']
        else:
            j['status'] = 'done'
            j['people'] = d.get('people') or []
    return jsonify({'ok': True})


@app.route('/ocr/status/<job_id>')
def ocr_status(job_id):
    """前端輪詢；完成後直接回傳比對結果"""
    with _ocr_lock:
        j = _ocr_jobs.get(job_id)
        if not j:
            return jsonify({'status': 'expired'}), 404
        st, people, err = j['status'], j['people'], j['error']
        af_bytes, af_name = j['af'], j['af_name']

    if st == 'need_review':
        return jsonify({'status': 'need_review_pending'})

    if st in ('pending', 'processing'):
        waited = 0
        with _ocr_lock:
            waited = int(time.time() - _ocr_jobs[job_id]['created'])
        if waited > 120:
            return jsonify({'status': 'timeout',
                            'error': '辨識服務目前無回應，請改用薪資系統直接匯出的 PDF。'})
        return jsonify({'status': st, 'waited': waited})

    if st == 'failed':
        return jsonify({'status': 'failed', 'error': err or '辨識失敗'})

    try:
        af_records, af_warns = paycheck.load_af(af_bytes, af_name)
        skey = school_key(af_name)
        ex = load_exclusions(skey)
        good, need = paycheck.from_ocr(people, af_records)
        af_names_set = {a['姓名'] for a in af_records}
        merge_title_stats(paycheck.title_observations(good + need, af_names_set))

        ud = ex.get('use_default', True)
        paycheck.mark_exclusions(good, ex['titles'], ex['names'], ud)
        paycheck.mark_exclusions(need, ex['titles'], ex['names'], ud)
        # 待確認清單中屬於「其他人員」者不必打擾使用者，直接沿用辨識值
        for n in [x for x in need if x.get('_次要')]:
            good.append({
                '姓名': (n.get('建議姓名') or n.get('原始姓名') or '').strip(),
                '職稱': n.get('職稱', ''),
                '薪俸': n.get('薪俸', 0), '專業加給': n.get('專業加給', 0),
                '主管加給': n.get('主管加給', 0), '導師特教': n.get('導師特教', 0),
                '應發金額': n.get('應發金額', 0),
                '_次要': True, '_次要原因': n.get('_次要原因', ''),
            })
        need = [n for n in need if not n.get('_次要')]

        if need:
            with _ocr_lock:
                if job_id in _ocr_jobs:
                    _ocr_jobs[job_id]['good'] = good
                    _ocr_jobs[job_id]['status'] = 'need_review'
            return jsonify({
                'status': 'need_review', 'success': True, 'job_id': job_id,
                'auto_ok': len(good), 'need': need, 'school_key': skey,
                'af_names': sorted({a['姓名'] for a in af_records}),
                'warnings': af_warns,
            })

        out = paycheck.compare(good, af_records)
        out.update({'status': 'done', 'success': True, 'mode': 'ocr',
                    'layout': '掃描圖檔（文字辨識）', 'af_count': len(af_records),
                    'pdf_count': len(good), 'warnings': af_warns,
                    'school_key': skey, 'counts': bump_counter('compares')})
        with _ocr_lock:
            _ocr_jobs.pop(job_id, None)
        return jsonify(out)
    except Exception as e:
        return jsonify({'status': 'failed', 'error': '比對錯誤：' + str(e)})


# ── 原有下載 / 列印路由 ────────────────────────────────────

@app.route('/download-simple')
def download_simple():
    if 'LAST_RESULT' not in app.config:
        return '尚無資料可下載', 400
    data = json.loads(app.config['LAST_RESULT'])
    all_cols = app.config['LAST_COLUMNS']
    cols = [c for c in SIMPLE_COLS if c in all_cols]
    out = build_excel(data, cols)
    return send_file(out, as_attachment=True, download_name='排序結果(簡單版).xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/download-simple-region')
def download_simple_region():
    if 'LAST_RESULT' not in app.config:
        return '尚無資料可下載', 400
    data = json.loads(app.config['LAST_RESULT'])
    all_cols = app.config['LAST_COLUMNS']
    cols = [c for c in SIMPLE_REGION_COLS if c in all_cols]
    out = build_excel(data, cols)
    return send_file(out, as_attachment=True, download_name='排序結果(簡單地域加給版).xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/download-result')
def download_result():
    if 'LAST_RESULT' not in app.config:
        return '尚無資料可下載', 400
    data = json.loads(app.config['LAST_RESULT'])
    columns = app.config['LAST_COLUMNS']
    out = build_excel(data, columns)
    return send_file(out, as_attachment=True, download_name='排序結果(完整版).xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _docx_to_pdf():
    school_name = app.config.get('LAST_SCHOOL', '')
    sn2 = app.config.get('LAST_SN2', '')
    yearmonth = app.config.get('LAST_YEARMONTH', '')
    try:
        docx_buf = build_audit_docx(school_name, sn2, yearmonth)
        docx_buf.seek(0)
        tmpdir = tempfile.mkdtemp()
        docx_path = os.path.join(tmpdir, 'audit.docx')
        pdf_path = os.path.join(tmpdir, 'audit.pdf')
        with open(docx_path, 'wb') as f:
            f.write(docx_buf.read())
        result = subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'pdf', '--outdir', tmpdir, docx_path],
            capture_output=True, timeout=60
        )
        if result.returncode != 0 or not os.path.exists(pdf_path):
            app.logger.error(f'LibreOffice error: {result.stderr}')
            return None
        with open(pdf_path, 'rb') as f:
            pdf_bytes = f.read()
        import shutil
        shutil.rmtree(tmpdir, ignore_errors=True)
        return pdf_bytes
    except Exception as e:
        app.logger.error(f'_docx_to_pdf exception: {e}')
        return None


def _build_audit_html(school_name, sn2, yearmonth, data):
    rows_html = ''
    for row in data:
        rows_html += f'''<tr>
      <td style="text-align:left">{row.get('姓名','')}</td>
      <td>{row.get('薪俸表別','')}</td><td>{row.get('支領數額','')}</td>
      <td>{row.get('專業加給表別','')}</td><td>{row.get('支領數額.1','')}</td>
      <td>{row.get('職務加給表別','')}</td><td>{row.get('支領數額.2','')}</td>
      <td>{row.get('地域加給表別','')}</td><td>{row.get('支領數額.3','')}</td>
    </tr>\n'''
    return f'''<h2 style="text-align:center;font-size:15px;margin-bottom:12px">高雄市政府教育局所屬機關學校 待遇稽核情形紀錄表</h2>
<div style="display:flex;gap:40px;margin-bottom:12px;font-size:13px">
  <span><strong style="color:#1a3a5c">學校名稱：</strong>{school_name or '（未對應）'}</span>
  <span><strong style="color:#1a3a5c">編號：</strong>{sn2 or '—'}</span>
  <span><strong style="color:#1a3a5c">稽核月份：</strong>{yearmonth or '—'}</span>
</div>
<table style="border-collapse:collapse;width:100%;font-size:12px">
  <thead>
    <tr>
      <th rowspan="2" style="border:1px solid #333;padding:6px 10px;background:#1a3a5c;color:#fff;-webkit-print-color-adjust:exact;print-color-adjust:exact">姓名</th>
      <th colspan="2" style="border:1px solid #333;padding:6px 10px;background:#1a3a5c;color:#fff;-webkit-print-color-adjust:exact;print-color-adjust:exact">薪俸</th>
      <th colspan="2" style="border:1px solid #333;padding:6px 10px;background:#1a3a5c;color:#fff;-webkit-print-color-adjust:exact;print-color-adjust:exact">專業加給</th>
      <th colspan="2" style="border:1px solid #333;padding:6px 10px;background:#1a3a5c;color:#fff;-webkit-print-color-adjust:exact;print-color-adjust:exact">職務加給</th>
      <th colspan="2" style="border:1px solid #333;padding:6px 10px;background:#1a3a5c;color:#fff;-webkit-print-color-adjust:exact;print-color-adjust:exact">地域加給</th>
    </tr>
    <tr>
      <th style="border:1px solid #333;padding:6px 8px;background:#1a3a5c;color:#fff;">表別</th>
      <th style="border:1px solid #333;padding:6px 8px;background:#1a3a5c;color:#fff;">支領數額</th>
      <th style="border:1px solid #333;padding:6px 8px;background:#1a3a5c;color:#fff;">表別</th>
      <th style="border:1px solid #333;padding:6px 8px;background:#1a3a5c;color:#fff;">支領數額</th>
      <th style="border:1px solid #333;padding:6px 8px;background:#1a3a5c;color:#fff;">表別</th>
      <th style="border:1px solid #333;padding:6px 8px;background:#1a3a5c;color:#fff;">支領數額</th>
      <th style="border:1px solid #333;padding:6px 8px;background:#1a3a5c;color:#fff;">表別</th>
      <th style="border:1px solid #333;padding:6px 8px;background:#1a3a5c;color:#fff;">支領數額</th>
    </tr>
  </thead>
  <tbody style="font-size:12px">{rows_html}</tbody>
</table>'''


@app.route('/print-simple')
def print_simple():
    if 'LAST_RESULT' not in app.config:
        return '尚無資料可列印', 400
    school_name = app.config.get('LAST_SCHOOL', '')
    data = json.loads(app.config.get('LAST_RESULT', '[]'))
    all_cols = app.config.get('LAST_COLUMNS', [])
    simple_cols = [c for c in SIMPLE_COLS if c in all_cols]

    thead = ''.join(f'<th>{c}</th>' for c in simple_cols)
    tbody = ''
    for row in data:
        cells = ''.join(
            f'<td{"" if c != "姓名" else " class=\"name\""} >{row.get(c, "")}</td>'
            for c in simple_cols
        )
        tbody += f'<tr>{cells}</tr>\n'

    return f'''<!DOCTYPE html>
<html lang="zh-TW"><head><meta charset="UTF-8">
<title>排序結果（簡單版）</title>
<style>
  body{{font-family:"Microsoft JhengHei",Arial,sans-serif;font-size:12px;margin:20px}}
  .school{{font-size:15px;font-weight:bold;color:#1a3a5c;margin-bottom:10px}}
  h3{{font-size:13px;color:#1a3a5c;margin-bottom:8px}}
  table{{border-collapse:collapse;width:100%}}
  th{{border:1px solid #333;padding:6px 10px;background:#1a3a5c;color:#fff;text-align:center;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
  td{{border:1px solid #aaa;padding:5px 8px;text-align:center}}
  td.name{{text-align:left;font-weight:600}}
  tr:nth-child(even){{background:#f0f4f9;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
  @media print{{body{{margin:8px}}}}
</style>
</head>
<body onload="window.print()">
<div class="school">{school_name}</div>
<h3>排序結果（簡單版）</h3>
<table><thead><tr>{thead}</tr></thead><tbody>{tbody}</tbody></table>
</body></html>'''


@app.route('/print-audit')
def print_audit():
    if 'LAST_SCHOOL' not in app.config:
        return 'No data', 400
    school_name = app.config.get('LAST_SCHOOL', '')
    sn2 = app.config.get('LAST_SN2', '')
    yearmonth = app.config.get('LAST_YEARMONTH', '')
    html = _build_audit_print_html(school_name, sn2, yearmonth, auto_print=True)
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


@app.route('/print-all')
def print_all():
    if 'LAST_RESULT' not in app.config:
        return 'No data', 400
    school_name = app.config.get('LAST_SCHOOL', '')
    sn2 = app.config.get('LAST_SN2', '')
    yearmonth = app.config.get('LAST_YEARMONTH', '')
    data = json.loads(app.config.get('LAST_RESULT', '[]'))
    all_cols = app.config.get('LAST_COLUMNS', [])
    simple_cols = [c for c in SIMPLE_COLS if c in all_cols]
    thead = ''.join('<th>' + c + '</th>' for c in simple_cols)
    tbody = ''
    for row in data:
        cells = ''
        for c in simple_cols:
            cls = ' class="name"' if c == '姓名' else ''
            cells += '<td' + cls + '>' + str(row.get(c, '')) + '</td>'
        tbody += '<tr>' + cells + '</tr>\n'
    audit_section = ''
    if school_name:
        audit_section = '<div class="audit-section">' + _build_audit_print_html(school_name, sn2, yearmonth, inner_only=True) + '</div>'
    page = (
        '<!DOCTYPE html><html lang="zh-TW"><head><meta charset="UTF-8">'
        '<title>' + school_name + '</title>'
        '<style>'
        'body{font-family:"Microsoft JhengHei",Arial,sans-serif;margin:20px;font-size:12px}'
        'h3{color:#1a3a5c;margin-bottom:8px;font-size:13px}'
        '.simple-table{border-collapse:collapse;width:100%;font-size:11px}'
        '.simple-table th{border:1px solid #333;padding:5px 8px;background:#1a3a5c;color:#fff;text-align:center;-webkit-print-color-adjust:exact;print-color-adjust:exact}'
        '.simple-table td{border:1px solid #aaa;padding:4px 8px;text-align:center}'
        '.simple-table td.name{text-align:left;font-weight:600}'
        '.simple-table tr:nth-child(even){background:#f0f4f9;-webkit-print-color-adjust:exact;print-color-adjust:exact}'
        '.school-name{font-size:14px;font-weight:700;color:#1a3a5c;margin-bottom:8px}'
        '.audit-section{page-break-before:right}'
        '.audit-tbl{border-collapse:collapse;width:100%;font-size:10px}'
        '.audit-tbl th,.audit-tbl td{border:1px solid #333;padding:4px 6px;text-align:center;vertical-align:middle}'
        '.audit-tbl th{background:#1a3a5c;color:#fff;-webkit-print-color-adjust:exact;print-color-adjust:exact}'
        '.audit-title{text-align:center;font-size:14px;font-weight:bold;margin:10px 0 8px}'
        '@media print{body{margin:8px}}'
        '</style></head>'
        '<body onload="window.print()">'
        '<div class="school-name">' + school_name + '</div>'
        '<h3>排序結果（簡單版）</h3>'
        '<table class="simple-table"><thead><tr>' + thead + '</tr></thead><tbody>' + tbody + '</tbody></table>'
        + audit_section +
        '</body></html>'
    )
    return page


def _build_audit_print_html(school_name, sn2, yearmonth, auto_print=False, inner_only=False):
    data_rows = ''
    for i in range(4):
        data_rows += (
            '<tr><td rowspan="2" style="text-align:left;min-width:60px"></td>'
            '<td>錯誤情形</td>'
            '<td></td><td colspan="2"></td><td colspan="2"></td><td></td>'
            '<td colspan="2"></td><td></td><td colspan="2"></td><td></td></tr>\n'
            '<tr><td>正確情形</td>'
            '<td></td><td colspan="2"></td><td colspan="2"></td><td></td>'
            '<td colspan="2"></td><td></td><td colspan="2"></td><td></td></tr>\n'
        )
    worker_rows = data_rows
    inner = (
        '<div class="audit-title">高雄市政府教育局所屬機關學校 待遇稽核情形紀錄表</div>'
        '<div style="font-size:11px;margin-bottom:8px">'
        '編號：' + (sn2 or '　　　') + '　　'
        '學校名稱：' + (school_name or '　　　　　　　') + '　　'
        '稽核月份：' + (yearmonth or '　　　') + '</div>'
        '<table class="audit-tbl">'
        '<thead>'
        '<tr>'
        '<th colspan="2">一般人員<br>(含約脩僱人員)稽核筆數</th>'
        '<th colspan="2"></th>'
        '<th colspan="2">錯誤筆數</th>'
        '<th colspan="3"></th>'
        '<th colspan="3">正確率</th>'
        '<th colspan="2">%</th>'
        '</tr>'
        '<tr>'
        '<th colspan="2" rowspan="2">姓　名</th>'
        '<th colspan="12">抽　驗　項　目</th>'
        '</tr>'
        '<tr>'
        '<th>薪俣表別</th>'
        '<th colspan="2">薪俣支領數額</th>'
        '<th colspan="2">專業加給表別</th>'
        '<th>專業加給支領數額</th>'
        '<th colspan="2">職務加給表別</th>'
        '<th>職務加給支領數額</th>'
        '<th colspan="2">地域加給表別</th>'
        '<th>地域加給支領數額</th>'
        '</tr>'
        '</thead>'
        '<tbody>' + data_rows + '</tbody>'
        '</table>'
        '<br>'
        '<table class="audit-tbl" style="margin-top:8px">'
        '<thead>'
        '<tr>'
        '<th colspan="2">技工工友稽核筆數</th>'
        '<th colspan="2"></th>'
        '<th colspan="2">錯誤筆數</th>'
        '<th colspan="3"></th>'
        '<th colspan="3">正確率</th>'
        '<th colspan="2">%</th>'
        '</tr>'
        '<tr>'
        '<th colspan="2" rowspan="2">姓　名</th>'
        '<th colspan="12">抽　驗　項　目</th>'
        '</tr>'
        '<tr>'
        '<th>薪俣表別</th>'
        '<th colspan="2">薪俣支領數額</th>'
        '<th colspan="2">專業加給表別</th>'
        '<th>專業加給支領數額</th>'
        '<th colspan="2">職務加給表別</th>'
        '<th>職務加給支領數額</th>'
        '<th colspan="2">地域加給表別</th>'
        '<th>地域加給支領數額</th>'
        '</tr>'
        '</thead>'
        '<tbody>' + worker_rows + '</tbody>'
        '</table>'
        '<div style="margin-top:16px;font-size:11px">'
        '稽核人員機關學校：　　　　　　　　　　　　職稱：　　　　　　姓名：＿＿＿＿＿＿＿<br>'
        '該組負責人核章：＿＿＿＿＿＿＿＿＿＿＿<br>'
        '本局承辦人：＿＿＿＿＿＿＿＿＿＿＿　　人事主任：＿＿＿＿＿＿＿＿＿＿＿＿＿'
        '</div>'
    )
    if inner_only:
        return inner
    onload = ' onload="window.print()"' if auto_print else ''
    return (
        '<!DOCTYPE html><html lang="zh-TW"><head><meta charset="UTF-8">'
        '<title>稽核表－' + school_name + '</title>'
        '<style>'
        'body{font-family:"Microsoft JhengHei",Arial,sans-serif;margin:20px;font-size:12px}'
        '.audit-title{text-align:center;font-size:14px;font-weight:bold;margin:10px 0 8px}'
        '.audit-tbl{border-collapse:collapse;width:100%;font-size:10px}'
        '.audit-tbl th,.audit-tbl td{border:1px solid #333;padding:4px 6px;text-align:center;vertical-align:middle}'
        '.audit-tbl th{background:#1a3a5c;color:#fff;-webkit-print-color-adjust:exact;print-color-adjust:exact}'
        '@media print{body{margin:8px}}'
        '</style></head>'
        '<body' + onload + '>'
        + inner +
        '</body></html>'
    )


@app.route('/download-audit')
def download_audit():
    if 'LAST_SCHOOL' not in app.config:
        return '尚無資料可下載', 400
    school_name = app.config.get('LAST_SCHOOL', '')
    sn2 = app.config.get('LAST_SN2', '')
    yearmonth = app.config.get('LAST_YEARMONTH', '')
    out = build_audit_docx(school_name, sn2, yearmonth)
    filename = f'稽核表_{school_name or "未知學校"}.docx'
    return send_file(out, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
